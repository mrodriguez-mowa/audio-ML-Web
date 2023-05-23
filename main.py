import re
from pydub import AudioSegment
import os
import requests as req
from pysentimiento import create_analyzer
import openpyxl

# Read the audio file
audio_list = os.listdir('./audios')
analyzer = create_analyzer(lang="es", task="sentiment")
report = []

analyzer_values = {
  "NEG": "Negativo",
  "NEU": "Neutro",
  "POS": "Positivo"
}

for audio in audio_list:
  
  current_audio = AudioSegment.from_file('./audios/' + audio, format="wav")
  if current_audio.duration_seconds <= 15:
    print('AUDIO MUY CORTO')
  else:
    print('SE PROCESARÁ POR WHISPER Y GENERARÁ UN ARCHIVO DE TEXTO')
    # SE LEE EL ARCHIVO DE TEXTO #
    transcript_file = f'./txt/{audio[:-4]}.txt'
    valid_transcript = []
    with open(transcript_file, 'r', encoding='utf-8') as file:
      transcript = file.read()
    
    # SE QUITAN LOS TIMESTAMPS Y SPEAKERS #
    new_transcript = ""
    text = transcript.split('\n')
    for line in text:
      if not line.upper().__contains__('SPEAKER'):
          valid_transcript.append(line)
      new_transcript = ' '.join(valid_transcript)
    
    if (len(new_transcript) <= 20):
      print('NO ES UN AUDIO VÁLIDO PARA PROCESAR')
    else:
      print('ES UN AUDIO VÁLIDO PARA PROCESAR')
      # SE HACE ANALIZA TODO EL TEXTO #
      # print(analyzer.predict(new_transcript))
      # SE SEPARARÁ TEXTO POR SPEAKER #
      value_report = {
        'audio': audio,
        'txt': audio[:-4] + '.txt',
        'speaker_1': '',
        'speaker_2': '',
      }
      speaker_1_text = []
      speaker_2_text = []
      current_speaker = 0
      for line in text:
        if line.upper().__contains__('SPEAKER'):
          id_val = line.split(' ')[1]
          current_speaker = int(id_val)
        else:
          if current_speaker == 1:
            speaker_1_text.append(line)
          else:
            speaker_2_text.append(line)

      # SE ANALIZA EL TEXTO DE CADA SPEAKER #
      speaker_1_text = ' '.join(speaker_1_text)
      speaker_2_text = ' '.join(speaker_2_text)

      speaker_1_sentiment = analyzer.predict(speaker_1_text)
      speaker_2_sentiment = analyzer.predict(speaker_2_text)

      

      value_report['speaker_1'] = analyzer_values[str(speaker_1_sentiment).replace("AnalyzerOutput(","").split(",")[0].split("=")[1]]
      value_report['speaker_2'] = analyzer_values[str(speaker_2_sentiment).replace("AnalyzerOutput(","").split(",")[0].split("=")[1]]

      report.append(value_report)

      # SE GENERA EL REPORTE EN EXCEL #
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = 'Reporte'

      ws['A1'] = 'Audio'
      ws['B1'] = 'Txt'
      ws['C1'] = 'Speaker 1'
      ws['D1'] = 'Speaker 2'

      for i in range(len(report)):
        ws[f'A{i + 2}'] = report[i]['audio']
        ws[f'B{i + 2}'] = report[i]['txt']
        ws[f'C{i + 2}'] = report[i]['speaker_1']
        ws[f'D{i + 2}'] = report[i]['speaker_2']

      wb.save('reporte.xlsx')
      




"""


# Read the transcript file
filename = './txt/transcript.txt'
with open(filename, 'r', encoding='utf-8') as file:
    transcript = file.read()
    
# Remove timestamps and speakers
valid_transcript = []

text = transcript.split('\n')

for line in text:
    if not line.upper().__contains__('SPEAKER'):
        valid_transcript.append(line)

new_transcript = ' '.join(valid_transcript)

# Count the number of words

words = new_transcript.split(' ')

if (len(words) <= 20):
  print('NO ES UN AUDIO VÁLIDO PARA PROCESAR')
else:
  print('ES UN AUDIO VÁLIDO PARA PROCESAR')


1. Validar Longitud de Audio
2. Procesar por whisper
3. Obtener el texto
4. Validar Longitud de texto
5. Procesar cada línea de texto por NLP
6. Obtener un promedio 
"""

def remove_timestamps_speakers(text):
    # Remove timestamps and speakers
    text = re.sub(r'\d{2}:\d{2}:\d{2}\s\w+\s', '', text)
    return text